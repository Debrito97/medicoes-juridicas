import streamlit as st
import pandas as pd
from datetime import datetime, date
import re
import io
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image as OpenPyXLImage
from openpyxl.workbook.protection import WorkbookProtection
import os
import base64

# ============================================================
# CONFIGURAÇÃO INICIAL E ESTADOS
# ============================================================

def initialize_session_state():
    """Inicializa todos os estados da sessão"""
    if 'is_dados_validado' not in st.session_state:
        st.session_state.is_dados_validado = False
    if 'is_revisao_concluida' not in st.session_state:
        st.session_state.is_revisao_concluida = False
    if 'is_detalhamento_iniciado' not in st.session_state:
        st.session_state.is_detalhamento_iniciado = False
    if 'is_detalhamento_validado' not in st.session_state:
        st.session_state.is_detalhamento_validado = False
    if 'is_finalizado' not in st.session_state:
        st.session_state.is_finalizado = False
    
    if 'dados_iniciais' not in st.session_state:
        st.session_state.dados_iniciais = {}
    if 'dados_coletados' not in st.session_state:
        st.session_state.dados_coletados = []
    if 'dados_cobrancas_nao_validados' not in st.session_state:
        st.session_state.dados_cobrancas_nao_validados = []
    if 'frames_cobrancas_data' not in st.session_state:
        st.session_state.frames_cobrancas_data = []
    
    if 'current_page' not in st.session_state:
        st.session_state.current_page = "inicio"
    
    if 'cobrancas' not in st.session_state:
        st.session_state.cobrancas = [{}]  # Lista de dicionários para cada cobrança

# ============================================================
# CONSTANTES E DICIONÁRIOS
# ============================================================

# Cores
PRIMARY_BLUE = "#00509d"
SIDEBAR_BLUE = "#3465a4"
SIDEBAR_HOVER = "#2b5285"
LIGHT_GRAY_BG = "#f4f4f4"
RESUMO_BLUE_BG = "#e8f4ff"
DISABLED_BG = "#eeeeee"
ERROR_RED_BG = "#FFE6E6"
RED_FG = "#C00000"
SUCCESS_GREEN = "#28a745"
WARNING_RED = "#dc3545"
WARNING_RED_HOVER = "#c82333"

# Dicionários de Mapeamento (De-Para)
MAPA_TIPO_COBRANCA = {
    "Consultorias": "CONS",
    "Despesas": "DESP",
    "Êxito": "EXITO",
    "Honorários Advocatícios (Prolabore)": "HON",
    "Parecer": "PARC",
    "Perícia": "PERI",
    "Publicações Legais": "PUB"
}

MAPA_MATERIA = {
    "4819": "4819",
    "Ambiental": "AMB",
    "Cálculos Trabalhistas": "CALCTR",
    "Cível": "CIV",
    "Contratos": "CONT",
    "Debêntures": "DEBS",
    "Fundiário": "FND",
    "Novos Negócios": "NNG",
    "Regulatório": "REG",
    "Publicações Legais": "PUBLEG",
    "BPO": "BPO",
    "Acompanhamento de Processos Judiciais": "ACPROC",
    "Recorte Eletrônico dos diários oficiais": "REDO",
    "Trabalhista": "TRB",
    "Tributário": "TRI",
    "Aduaneiro": "ADU",
    "Levantamento de Processos": "LEVPROC",
    "Societário": "SOC"
}

# Listas de valores
MATERIAS_JURIDICAS = [
    "4819", "Ambiental", "Cálculos Trabalhistas", "Cível", "Contratos",
    "Debêntures", "Fundiário", "Novos Negócios", "Regulatório", 
    "Publicações Legais", "BPO", "Acompanhamento de Processos Judiciais",
    "Recorte Eletrônico dos diários oficiais", "Trabalhista", "Tributário",
    "Aduaneiro", "Levantamento de Processos", "Societário"
]

TIPOS_COBRANCA = [
    "Consultorias", "Despesas", "Êxito", "Honorários Advocatícios (Prolabore)",
    "Parecer", "Perícia", "Publicações Legais"
]

# Dados dos projetos por empresa (ATUALIZADO - CAMPOS VAZIOS)
PROJETOS_POR_EMPRESA = {
    "Interligação Elétrica Aguapeí": {
        "Projeto vinculado": ["IE Interligação Elétrica Aguapeí"],
        "Trecho": {"IE Interligação Elétrica Aguapeí": [""]}
    },
    "Interligação Elétrica Biguaçu": {
        "Projeto vinculado": ["IE Interligação Elétrica Biguaçu"],
        "Trecho": {"IE Interligação Elétrica Biguaçu": [""]}
    },
    "Interligação Elétrica Evrecy": {
        "Projeto vinculado": ["Minuano"],
        "Trecho": {"Minuano": [""]}
    },
    "ISA ENERGIA BRASIL": {
        "Projeto vinculado": ["Replan", "Fernão Dias", "Piraque", "Itatiaia", "Serra Dourada"],
        "Trecho": {
            "Replan": [""],
            "Fernão Dias": [""],
            "Piraque": [
                "LT 500 kv JAIBA-JANAUBA6", "LT 500 Kv JANAUBA6-JANAUBA3",
                "LT 500 kV JANAUBA6 -CAPELINHA 3", "LT 500 kV CAPELINHA 3 -GOVERNADOR VALADARES",
                "LT 500 kV João Neiv 2 - Viana 2 C1", "LT 345 kV Viana 2 - Viana, C3",
                "SE Janaúba 6", "SE Capelinha 3", "SE Jaíba", "SE Janaúba 3",
                "SE Governador Valadares 6", "SE João Neiva 2", "SE Viana 2", "SE Viana"
            ],
            "Itatiaia": ["Governador Valadares 6 - Leopoldina 2", "Leopoldina 2 - Terminal Rio"],
            "Serra Dourada": [
                "SE BURITIRAMA", "SE BARRA II", "SE CORRENTINA", "SE ARINOS", "SE CAMPO FORMOSO",
                "SE JUAZEIRO", "SE BOM JESUS DA LAPA", "SE RIO DAS ÉGUAS",
                "LT 500 kV Buritirama -Barra II C1, CS", "LT 500 kV Barra II -Correntina C1, CS",
                "LT 500 kV Correntina -Arinos 2 C1, CS", "LT 500 kV Juazeiro III -Campo Formoso II",
                "LT 500kV Campo Formoso II Barra II C1 C2", "LT 500kV Bom Jesus da Lapa- Rio das Égua"
            ]
        }
    },
    "Interligação Elétrica Itaúnas": {
        "Projeto vinculado": ["IE Interligação Elétrica Itaúnas"],
        "Trecho": {"IE Interligação Elétrica Itaúnas": [""]}
    },
    "Interligação Elétrica Ivaí": {
        "Projeto vinculado": ["IE Interligação Elétrica Ivaí"],
        "Trecho": {
            "IE Interligação Elétrica Ivaí": [
                "ERB - LOTE 1", "ODI - SE Guaíra", "ODI - SE Sarandi", "ODI - SE Paranavaí Norte",
                "ODI - SE Foz do Iguaçu", "ODI - SE Londrina", "ODI - LT Guaíra - Sarandi",
                "ODI - LT Foz do Iguaçu - Guaíra", "ODI - LT Londrina - Sarandi",
                "ODI - LT Sarandi - Paranavaí Norte", "ODI - Administração - Técnica",
                "ODI - Adiantamento a Fornecedor"
            ]
        }
    },
    "Interligação Elétrica Jaguar 8": {
        "Projeto vinculado": ["Água Azul"],
        "Trecho": {"Água Azul": [""]}
    },
    "Interligação Elétrica Minas Gerais": {
        "Projeto vinculado": ["IEMG", "Triângulo Mineiro"],
        "Trecho": {"IEMG": [""], "Triângulo Mineiro": [""]}
    },
    "Interligação Elétrica Norte Nordeste": {
        "Projeto vinculado": ["IENNE"],
        "Trecho": {"IENNE": [""]}
    },
    "Interligação Elétrica Riacho Grande": {
        "Projeto vinculado": ["IE Interligação Elétrica Riacho Grande"],
        "Trecho": {"IE Interligação Elétrica Riacho Grande": [""]}
    },
    "Interligação Elétrica Sul": {
        "Projeto vinculado": ["IESUL"],
        "Trecho": {"IESUL": [""]}
    },
    "Interligação Elétrica Tibagi": {
        "Projeto vinculado": ["IE Interligação Elétrica Tibagi"],
        "Trecho": {"IE Interligação Elétrica Tibagi": [""]}
    }
}

# Listas para comboboxes
EMPRESAS = [
    "ISA ENERGIA BRASIL", "Interligação Elétrica Evrecy", "Interligação Elétrica Minas Gerais",
    "Interligação Elétrica Norte Nordeste", "Interligação Elétrica Pinheiros",
    "Interligação Elétrica Sul", "Interligação Elétrica Serra Japi", "Interligação Elétrica Itaúnas",
    "Interligação Elétrica Itapura", "Interligação Elétrica Aguapeí", "Interligação Elétrica Itaquerê",
    "Interligação Elétrica Tibagi", "Interligação Elétrica Ivaí", "Interligação Elétrica Biguaçu",
    "Interligação Elétrica Jaguar 6", "Interligação Elétrica Jaguar 8", "Interligação Elétrica Jaguar 9",
    "Interligação Elétrica Riacho Grande",
]

ADVOGADOS = [
    "Andrea Mazzaro Carlos de Vincenti", "Carlos Lopes", "Emerson Rodrigues do Nascimento",
    "Eric Tadao Pagani Fukai", "Erica Barbeiro Travassos", "Francisco Ricardo Tavian",
    "Gilvan Aparecido dos Santos", "Leonam Ricardo Alcantara Francisconi",
    "Leonardo Lupercio Garcia Martins", "Leonardo Silva Merces", "Letticia Pinheiro de Oliveira Barros",
    "Luciana Semenzato Garcia", "Marjorie Merida Chiesa", "Natalia Mendonca Goncalves",
    "Pedro Henrique Ribeiro e Silva", "Ricardo de Oliveira Beninca", "Rita Halabian",
]

TIPOS_DOCUMENTO = ["Nota Fiscal", "Nota de Débito", "RPA (Recibo de Pagamento Autônomo)"]

# ============================================================
# FUNÇÕES DE APOIO
# ============================================================

def filtrar_numeros(texto: str) -> str:
    """Remove todos os caracteres não numéricos de uma string."""
    return "".join(ch for ch in texto if ch.isdigit())

def cnpj_valido(cnpj: str) -> bool:
    """Verifica se o CNPJ é válido (simplificado)."""
    if len(cnpj) != 14 or cnpj == cnpj[0] * 14:
        return False

    pesos1 = [5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]
    pesos2 = [6, 5, 4, 3, 2, 9, 8, 7, 6, 5, 4, 3, 2]

    # Cálculo do primeiro dígito verificador
    soma = sum(int(cnpj[i]) * pesos1[i] for i in range(12))
    resto = soma % 11
    digito1 = 0 if resto < 2 else 11 - resto

    # Cálculo do segundo dígito verificador
    soma = sum(int(cnpj[i]) * pesos2[i] for i in range(13))
    resto = soma % 11
    digito2 = 0 if resto < 2 else 11 - resto

    return cnpj[12] == str(digito1) and cnpj[13] == str(digito2)

def formatar_cnpj(cnpj: str) -> str:
    """Formata CNPJ para o padrão XX.XXX.XXX/XXXX-XX"""
    cnpj_numeros = filtrar_numeros(cnpj)
    if len(cnpj_numeros) != 14:
        return cnpj_numeros
    
    return f"{cnpj_numeros[:2]}.{cnpj_numeros[2:5]}.{cnpj_numeros[5:8]}/{cnpj_numeros[8:12]}-{cnpj_numeros[12:14]}"

def formatar_moeda(valor: str) -> str:
    """Formata string de moeda para o padrão brasileiro"""
    texto_puro = re.sub(r'[^0-9]', '', valor)
    
    if not texto_puro:
        return ""
    
    if len(texto_puro) < 3:
        texto_puro = '0' * (3 - len(texto_puro)) + texto_puro

    inteiro = texto_puro[:-2]
    decimal = texto_puro[-2:]

    if not inteiro: 
        inteiro = "0"

    try:
        inteiro_num = int(inteiro)
        inteiro = str(inteiro_num)
    except ValueError:
        inteiro = "0"

    inteiro_formatado = f"{int(inteiro):,}".replace(",", ".")
    return f"{inteiro_formatado},{decimal}"

def validar_contrato(contrato: str) -> bool:
    """Valida o formato do contrato"""
    if len(contrato) != 10:
        return False
    return True

def validar_pedido(pedido: str) -> bool:
    """Valida o formato do pedido"""
    if len(pedido) != 10:
        return False
    return pedido.isdigit()

def atualizar_texto_breve(tipo_cobranca: str, materia: str) -> str:
    """Gera o texto breve baseado no tipo de cobrança e matéria"""
    if not tipo_cobranca or not materia:
        return "Aguardando seleção..."
    
    sigla_tipo = MAPA_TIPO_COBRANCA.get(tipo_cobranca, "TIPO?")
    sigla_materia = MAPA_MATERIA.get(materia, "MAT?")
    return f"{sigla_tipo}_{sigla_materia}"

# ============================================================
# FUNÇÕES DE NAVEGAÇÃO
# ============================================================

def navegar_para(pagina: str):
    """Navega para uma página específica"""
    st.session_state.current_page = pagina
    st.rerun()

def verificar_acesso(pagina: str) -> bool:
    """Verifica se o usuário pode acessar a página baseado no fluxo"""
    if pagina == "revisao" and not st.session_state.is_dados_validado:
        st.error("Você deve preencher e validar os Dados Iniciais antes de avançar para a Revisão.")
        return False
    elif pagina == "detalhamento" and not st.session_state.is_revisao_concluida:
        st.error("Você deve confirmar a Revisão dos Dados Iniciais antes de avançar para o Detalhamento.")
        return False
    elif pagina == "revisao_detalhada" and not st.session_state.is_detalhamento_validado:
        st.error("Ocorreu um erro. A validação do detalhamento não foi concluída.")
        return False
    return True

# ============================================================
# PÁGINAS DO SISTEMA
# ============================================================

def pagina_inicio():
    """Página inicial do sistema"""
    st.markdown(
        f"""
        <div style="text-align: center; padding: 50px; background-color: white; border-radius: 10px;">
            <h1 style="color: {PRIMARY_BLUE}; font-size: 2.5em;">Sistema de Medições Jurídicas</h1>
            <h3 style="color: #555555;">Módulo de Cadastro Inicial</h3>
            <p style="color: #666666; font-size: 1.1em;">(Dados para Medições Jurídicas)</p>
        </div>
        """, 
        unsafe_allow_html=True
    )
    
    st.write("")
    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        if st.button("Iniciar Lançamento", type="primary", use_container_width=True):
            navegar_para("dados")

def pagina_dados():
    """Página de dados iniciais"""
    st.header("Dados Iniciais - Faturamento")
    
    with st.form("dados_iniciais"):
        col1, col2 = st.columns(2)
        
        with col1:
            # CNPJ
            cnpj = st.text_input("CNPJ do fornecedor*", 
                               value=st.session_state.dados_iniciais.get('cnpj', ''),
                               help="Digite o CNPJ no formato XX.XXX.XXX/XXXX-XX")
            
            # Empresa
            empresa = st.selectbox("Empresa contratante*", 
                                 options=[""] + sorted(EMPRESAS),
                                 index=0 if not st.session_state.dados_iniciais.get('empresa') 
                                 else sorted(EMPRESAS).index(st.session_state.dados_iniciais.get('empresa')) + 1)
            
            # Advogado
            advogado = st.selectbox("Advogado (a) responsável*", 
                                  options=[""] + sorted(ADVOGADOS),
                                  index=0 if not st.session_state.dados_iniciais.get('advogado')
                                  else sorted(ADVOGADOS).index(st.session_state.dados_iniciais.get('advogado')) + 1)
            
            # Tipo de documento
            tipo_doc = st.selectbox("Tipo de documento de cobrança*", 
                                  options=[""] + TIPOS_DOCUMENTO,
                                  index=0 if not st.session_state.dados_iniciais.get('tipo_doc')
                                  else TIPOS_DOCUMENTO.index(st.session_state.dados_iniciais.get('tipo_doc')) + 1)
        
        with col2:
            # Data prevista
            data_prevista = st.date_input("Data prevista de emissão*", 
                                        value=st.session_state.dados_iniciais.get('data_prevista', datetime.now().date()))
            
            # Contrato vinculado
            existe_contrato = st.radio("Existe contrato vinculado?", 
                                     options=["Não", "Sim"],
                                     index=0 if st.session_state.dados_iniciais.get('existe_contrato', 'Não') == "Não" else 1,
                                     horizontal=True)
            
            if existe_contrato == "Sim":
                n_contrato = st.text_input("Nº do contrato*", 
                                         value=st.session_state.dados_iniciais.get('n_contrato', ''),
                                         max_chars=10,
                                         help="Formato: 2 caracteres alfanuméricos + 8 numéricos")
            else:
                n_contrato = ""
            
            # Pedido vinculado
            existe_pedido = st.radio("Existe pedido vinculado?", 
                                   options=["Não", "Sim"],
                                   index=0 if st.session_state.dados_iniciais.get('existe_pedido', 'Não') == "Não" else 1,
                                   horizontal=True)
            
            if existe_pedido == "Sim":
                n_pedido = st.text_input("Nº do pedido*", 
                                       value=st.session_state.dados_iniciais.get('n_pedido', ''),
                                       max_chars=10,
                                       help="10 caracteres numéricos")
            else:
                n_pedido = ""
        
        # Campos full width
        n_medicao = st.text_input("Nº medição (doc.interno fornecedor)*",
                                value=st.session_state.dados_iniciais.get('n_medicao', ''))
        
        breve_desc = st.text_area("Breve descrição da fatura*",
                                value=st.session_state.dados_iniciais.get('breve_desc', ''),
                                height=80)
        
        # Botão de validação
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            if st.form_submit_button("Validar Dados e Prosseguir para Revisão", type="primary", use_container_width=True):
                # Validações
                erros = []
                
                # CNPJ
                cnpj_numeros = filtrar_numeros(cnpj)
                if len(cnpj_numeros) == 0:
                    erros.append("CNPJ é obrigatório")
                elif len(cnpj_numeros) != 14:
                    erros.append("CNPJ incompleto. Verifique o número digitado.")
                elif not cnpj_valido(cnpj_numeros):
                    erros.append("CNPJ inválido. Verifique o número digitado.")
                
                # Empresa
                if not empresa:
                    erros.append("Empresa contratante é obrigatória")
                
                # Advogado
                if not advogado:
                    erros.append("Advogado responsável é obrigatório")
                
                # Tipo documento
                if not tipo_doc:
                    erros.append("Tipo de documento é obrigatório")
                
                # Contrato
                if existe_contrato == "Sim" and not n_contrato:
                    erros.append("Nº do contrato é obrigatório quando 'Sim' é selecionado")
                elif existe_contrato == "Sim" and not validar_contrato(n_contrato):
                    erros.append("Nº do contrato deve ter 10 caracteres no formato correto")
                
                # Pedido
                if existe_pedido == "Sim" and not n_pedido:
                    erros.append("Nº do pedido é obrigatório quando 'Sim' é selecionado")
                elif existe_pedido == "Sim" and not validar_pedido(n_pedido):
                    erros.append("Nº do pedido deve ter 10 caracteres numéricos")
                
                # Nº medição
                if not n_medicao:
                    erros.append("Nº medição é obrigatório")
                
                # Breve descrição
                if not breve_desc:
                    erros.append("Breve descrição é obrigatória")
                
                if erros:
                    for erro in erros:
                        st.error(erro)
                else:
                    # Salvar dados
                    st.session_state.dados_iniciais = {
                        "cnpj": formatar_cnpj(cnpj),
                        "empresa": empresa,
                        "advogado": advogado,
                        "tipo_doc": tipo_doc,
                        "data_prevista": data_prevista,
                        "existe_contrato": existe_contrato,
                        "n_contrato": n_contrato,
                        "existe_pedido": existe_pedido,
                        "n_pedido": n_pedido,
                        "n_medicao": n_medicao,
                        "breve_desc": breve_desc
                    }
                    st.session_state.is_dados_validado = True
                    st.success("Dados validados com sucesso! Avançando para revisão...")
                    st.rerun()

def pagina_revisao():
    """Página de revisão dos dados iniciais"""
    if not verificar_acesso("revisao"):
        return
    
    st.header("Revisão dos Dados Iniciais")
    
    st.info("Por favor, revise os dados antes de prosseguir. Estes são os dados principais que serão usados para todas as cobranças.")
    
    # Resumo dos dados
    st.subheader("Resumo dos Dados Iniciais")
    
    dados = st.session_state.dados_iniciais
    col1, col2 = st.columns(2)
    
    with col1:
        st.write(f"**CNPJ Fornecedor:** {dados.get('cnpj', 'N/A')}")
        st.write(f"**Empresa Contratante:** {dados.get('empresa', 'N/A')}")
        st.write(f"**Advogado (a) responsável:** {dados.get('advogado', 'N/A')}")
        st.write(f"**Tipo de Documento:** {dados.get('tipo_doc', 'N/A')}")
        st.write(f"**Data de Emissão:** {dados.get('data_prevista').strftime('%d/%m/%Y') if dados.get('data_prevista') else 'N/A'}")
    
    with col2:
        # Contrato
        existe_contrato = dados.get('existe_contrato', 'Não')
        if existe_contrato == "Sim":
            texto_contrato = f"Sim (Nº {dados.get('n_contrato', 'N/A')})"
        else:
            texto_contrato = "Não"
        st.write(f"**Contrato Vinculado:** {texto_contrato}")
        
        # Pedido
        existe_pedido = dados.get('existe_pedido', 'Não')
        if existe_pedido == "Sim":
            texto_pedido = f"Sim (Nº {dados.get('n_pedido', 'N/A')})"
        else:
            texto_pedido = "Não"
        st.write(f"**Pedido Vinculado:** {texto_pedido}")
        
        st.write(f"**Nº medição:** {dados.get('n_medicao', 'N/A')}")
        st.write(f"**Breve descrição:** {dados.get('breve_desc', 'N/A')}")
    
    # Botões de ação
    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        if st.button("← Voltar e Corrigir", use_container_width=True):
            navegar_para("dados")
    with col3:
        if st.button("Iniciar Detalhamento →", type="primary", use_container_width=True):
            st.session_state.is_revisao_concluida = True
            st.session_state.cobrancas = [{}]  # Inicializar lista de cobranças
            st.success("Revisão confirmada! Avançando para detalhamento...")
            st.rerun()

def pagina_detalhamento():
    """Página de detalhamento das cobranças"""
    if not verificar_acesso("detalhamento"):
        return
    
    st.header("Detalhamento das Cobranças")
    st.write("Preencha as informações de cada cobrança nos campos abaixo.")
    
    # Inicializar cobranças se necessário
    if 'cobrancas' not in st.session_state:
        st.session_state.cobrancas = [{}]
    
    # Formulário de cobranças
    for i, cobranca in enumerate(st.session_state.cobrancas):
        with st.expander(f"Cobrança {i+1}", expanded=True):
            render_cobranca_form(i, cobranca)
    
    # Botões para gerenciar cobranças
    col1, col2, col3 = st.columns([1, 1, 1])
    with col1:
        if st.button("+ Adicionar Nova Cobrança"):
            st.session_state.cobrancas.append({})
            st.rerun()
    
    # Botões de navegação
    st.markdown("---")
    col1, col2, col3, col4 = st.columns([1, 1, 1, 1])
    
    with col1:
        if st.button("Visualizar Resumo"):
            mostrar_resumo_popup()
    
    with col3:
        if st.button("← Voltar"):
            navegar_para("revisao")
    
    with col4:
        if st.button("Validar e Revisar →", type="primary"):
            if validar_detalhamento():
                st.session_state.is_detalhamento_validado = True
                st.success("Detalhamento validado com sucesso! Avançando para revisão detalhada...")
                st.rerun()

def render_cobranca_form(index, cobranca_data):
    """Renderiza o formulário para uma cobrança específica"""
    # Chaves para session_state
    prefix = f"cobranca_{index}"
    
    col1, col2 = st.columns(2)
    
    with col1:
        # Possui Nº Espaider?
        possui_espaider = st.radio(
            "Possui Nº Espaider?",
            options=["Não", "Sim"],
            key=f"{prefix}_possui_espaider",
            horizontal=True,
            index=0 if cobranca_data.get('possui_espaider', 'Não') == 'Não' else 1
        )
        
        # Nº Espaider (condicional)
        if possui_espaider == "Sim":
            n_espaider = st.text_input(
                "Nº Espaider*",
                value=cobranca_data.get('n_espaider', ''),
                key=f"{prefix}_n_espaider"
            )
        else:
            n_espaider = ""
        
        # Matéria
        materia = st.selectbox(
            "Matéria*",
            options=[""] + MATERIAS_JURIDICAS,
            key=f"{prefix}_materia",
            index=0 if not cobranca_data.get('materia') else MATERIAS_JURIDICAS.index(cobranca_data.get('materia')) + 1
        )
    
    with col2:
        # Possui projeto vinculado?
        possui_projeto = st.radio(
            "Possui projeto vinculado?",
            options=["Não", "Sim"],
            key=f"{prefix}_possui_projeto",
            horizontal=True,
            index=0 if cobranca_data.get('possui_projeto', 'Não') == 'Não' else 1
        )
        
        # Projeto e Trecho (condicionais)
        if possui_projeto == "Sim":
            empresa = st.session_state.dados_iniciais.get('empresa', '')
            projetos = PROJETOS_POR_EMPRESA.get(empresa, {}).get("Projeto vinculado", [""])
            
            projeto = st.selectbox(
                "Projeto vinculado*",
                options=projetos,
                key=f"{prefix}_projeto"
            )
            
            if projeto:
                trechos = PROJETOS_POR_EMPRESA[empresa]["Trecho"].get(projeto, [""])
                trecho = st.selectbox(
                    "Trecho",
                    options=trechos,
                    key=f"{prefix}_trecho"
                )
            else:
                trecho = ""
        else:
            projeto = ""
            trecho = ""
    
    # Mais de uma cobrança?
    mais_cobrancas = st.radio(
        "Existe mais de uma cobrança para este Nº Espaider?",
        options=["Não", "Sim"],
        key=f"{prefix}_mais_cobrancas",
        horizontal=True,
        index=0 if cobranca_data.get('mais_cobrancas', 'Não') == 'Não' else 1
    )
    
    st.markdown("---")
    
    # Cobrança Principal
    st.subheader("Cobrança Principal")
    
    col1, col2, col3 = st.columns([2, 1, 2])
    
    with col1:
        tipo_1 = st.selectbox(
            "Tipo de cobrança*",
            options=[""] + TIPOS_COBRANCA,
            key=f"{prefix}_tipo_1",
            index=0 if not cobranca_data.get('tipo_1') else TIPOS_COBRANCA.index(cobranca_data.get('tipo_1')) + 1
        )
    
    with col2:
        valor_1 = st.text_input(
            "Valor (R$)*",
            value=cobranca_data.get('valor_1', '0,00'),
            key=f"{prefix}_valor_1"
        )
    
    with col3:
        # CORREÇÃO: Texto breve atualizado automaticamente
        # Obter valores atuais dos campos
        current_tipo_1 = st.session_state.get(f"{prefix}_tipo_1", "")
        current_materia = st.session_state.get(f"{prefix}_materia", "")
        
        # Calcular texto breve automaticamente
        if current_tipo_1 and current_materia:
            texto_breve_1 = atualizar_texto_breve(current_tipo_1, current_materia)
        else:
            texto_breve_1 = "Aguardando seleção..."
        
        st.text_input(
            "Texto Breve Código Serviço",
            value=texto_breve_1,
            key=f"{prefix}_texto_1",
            disabled=True
        )
    
    # Cobrança Secundária (condicional)
    if mais_cobrancas == "Sim":
        st.subheader("Cobrança Secundária")
        
        col1, col2, col3 = st.columns([2, 1, 2])
        
        with col1:
            tipo_2 = st.selectbox(
                "Tipo de cobrança*",
                options=[""] + TIPOS_COBRANCA,
                key=f"{prefix}_tipo_2",
                index=0 if not cobranca_data.get('tipo_2') else TIPOS_COBRANCA.index(cobranca_data.get('tipo_2')) + 1
            )
        
        with col2:
            valor_2 = st.text_input(
                "Valor (R$)*",
                value=cobranca_data.get('valor_2', '0,00'),
                key=f"{prefix}_valor_2"
            )
        
        with col3:
            # CORREÇÃO: Texto breve atualizado automaticamente para cobrança secundária
            current_tipo_2 = st.session_state.get(f"{prefix}_tipo_2", "")
            current_materia = st.session_state.get(f"{prefix}_materia", "")
            
            if current_tipo_2 and current_materia:
                texto_breve_2 = atualizar_texto_breve(current_tipo_2, current_materia)
            else:
                texto_breve_2 = "Aguardando seleção..."
            
            st.text_input(
                "Texto Breve Código Serviço",
                value=texto_breve_2,
                key=f"{prefix}_texto_2",
                disabled=True
            )
    
    # Botão para remover cobrança (exceto a primeira)
    if index > 0:
        if st.button(f"Remover Cobrança {index+1}", type="secondary"):
            st.session_state.cobrancas.pop(index)
            st.rerun()
    
    # Atualizar dados da cobrança com os valores CORRETOS do texto breve
    current_tipo_1 = st.session_state.get(f"{prefix}_tipo_1", "")
    current_materia = st.session_state.get(f"{prefix}_materia", "")
    current_tipo_2 = st.session_state.get(f"{prefix}_tipo_2", "")
    
    # Recalcular textos breves para salvar nos dados
    texto_breve_1_save = atualizar_texto_breve(current_tipo_1, current_materia) if current_tipo_1 and current_materia else "Aguardando seleção..."
    texto_breve_2_save = atualizar_texto_breve(current_tipo_2, current_materia) if current_tipo_2 and current_materia else "Aguardando seleção..."
    
    cobranca_atualizada = {
        'possui_espaider': possui_espaider,
        'n_espaider': n_espaider,
        'possui_projeto': possui_projeto,
        'projeto': projeto,
        'trecho': trecho,
        'materia': current_materia,
        'mais_cobrancas': mais_cobrancas,
        'tipo_1': current_tipo_1,
        'valor_1': valor_1,
        'texto_1': texto_breve_1_save,
        'tipo_2': current_tipo_2 if mais_cobrancas == "Sim" else "",
        'valor_2': valor_2 if mais_cobrancas == "Sim" else "",
        'texto_2': texto_breve_2_save if mais_cobrancas == "Sim" else ""
    }
    
    st.session_state.cobrancas[index] = cobranca_atualizada
def validar_detalhamento() -> bool:
    """Valida todo o detalhamento das cobranças"""
    erros = []
    
    for i, cobranca in enumerate(st.session_state.cobrancas):
        num = i + 1
        
        # Validar Nº Espaider se necessário
        if cobranca.get('possui_espaider') == "Sim" and not cobranca.get('n_espaider'):
            erros.append(f"Cobrança {num}: Nº Espaider é obrigatório quando 'Sim' é selecionado")
        
        # Validar projeto se necessário
        if cobranca.get('possui_projeto') == "Sim" and not cobranca.get('projeto'):
            erros.append(f"Cobrança {num}: Projeto vinculado é obrigatório quando 'Sim' é selecionado")
        
        # Validar matéria
        if not cobranca.get('materia'):
            erros.append(f"Cobrança {num}: Matéria é obrigatória")
        
        # Validar cobrança principal
        if not cobranca.get('tipo_1'):
            erros.append(f"Cobrança {num}: Tipo de cobrança (principal) é obrigatório")
        
        # Validar valor principal
        valor_1_limpo = re.sub(r'[^0-9]', '', cobranca.get('valor_1', ''))
        if not valor_1_limpo or int(valor_1_limpo) == 0:
            erros.append(f"Cobrança {num}: Valor (principal) deve ser maior que zero")
        
        # Validar texto breve principal
        if cobranca.get('texto_1') == "Aguardando seleção..." or not cobranca.get('texto_1'):
            erros.append(f"Cobrança {num}: Texto Breve não foi gerado automaticamente")
        
        # Validar cobrança secundária se necessário
        if cobranca.get('mais_cobrancas') == "Sim":
            if not cobranca.get('tipo_2'):
                erros.append(f"Cobrança {num}: Tipo de cobrança (secundário) é obrigatório")
            
            valor_2_limpo = re.sub(r'[^0-9]', '', cobranca.get('valor_2', ''))
            if not valor_2_limpo or int(valor_2_limpo) == 0:
                erros.append(f"Cobrança {num}: Valor (secundário) deve ser maior que zero")
            
            if cobranca.get('texto_2') == "Aguardando seleção..." or not cobranca.get('texto_2'):
                erros.append(f"Cobrança {num}: Texto Breve (secundário) não foi gerado automaticamente")
    
    if erros:
        for erro in erros:
            st.error(erro)
        return False
    
    # Coletar dados validados
    st.session_state.dados_coletados = []
    for i, cobranca in enumerate(st.session_state.cobrancas):
        dados_cobranca = {
            'num_cobranca': i + 1,
            'Possui Nº Espaider?': cobranca.get('possui_espaider'),
            'Nº Espaider': cobranca.get('n_espaider', ''),
            'Possui projeto vinculado?': cobranca.get('possui_projeto'),
            'Projeto vinculado': cobranca.get('projeto', ''),
            'Trecho': cobranca.get('trecho', ''),
            'Matéria': cobranca.get('materia'),
            'sim_nao': cobranca.get('mais_cobrancas'),
            'bloco_1': {
                "tipo": cobranca.get('tipo_1'),
                "materia": cobranca.get('materia'),
                "valor": cobranca.get('valor_1'),
                "texto_breve": cobranca.get('texto_1')
            }
        }
        
        if cobranca.get('mais_cobrancas') == "Sim":
            dados_cobranca['bloco_2'] = {
                "tipo": cobranca.get('tipo_2'),
                "materia": cobranca.get('materia'),
                "valor": cobranca.get('valor_2'),
                "texto_breve": cobranca.get('texto_2')
            }
        
        st.session_state.dados_coletados.append(dados_cobranca)
    
    st.session_state.dados_cobrancas_nao_validados = st.session_state.dados_coletados.copy()
    return True

def mostrar_resumo_popup():
    """Mostra resumo dos dados iniciais"""
    if not st.session_state.dados_iniciais:
        st.warning("Nenhum dado inicial foi validado ainda.")
        return
    
    dados = st.session_state.dados_iniciais
    
    # Contrato
    existe_contrato = dados.get('existe_contrato', 'Não')
    if existe_contrato == "Sim":
        texto_contrato = f"Sim (Nº {dados.get('n_contrato', 'N/A')})"
    else:
        texto_contrato = "Não"
    
    # Pedido
    existe_pedido = dados.get('existe_pedido', 'Não')
    if existe_pedido == "Sim":
        texto_pedido = f"Sim (Nº {dados.get('n_pedido', 'N/A')})"
    else:
        texto_pedido = "Não"
    
    resumo = f"""
    **Estes são os dados fixos para todas as cobranças:**
    
    • **CNPJ:** {dados.get('cnpj', 'N/A')}
    • **Empresa:** {dados.get('empresa', 'N/A')}
    • **Advogado(a):** {dados.get('advogado', 'N/A')}
    • **Documento:** {dados.get('tipo_doc', 'N/A')}
    • **Data Emissão:** {dados.get('data_prevista').strftime('%d/%m/%Y') if dados.get('data_prevista') else 'N/A'}
    • **Nº medição:** {dados.get('n_medicao', 'N/A')}
    • **Breve descrição:** {dados.get('breve_desc', 'N/A')}
    • **Contrato:** {texto_contrato}
    • **Pedido:** {texto_pedido}
    """
    
    st.info(resumo)

def pagina_revisao_detalhada():
    """Página de revisão detalhada"""
    if not verificar_acesso("revisao_detalhada"):
        return
    
    st.header("Revisão do Detalhamento")
    
    st.info("Confirme os dados de todas as cobranças. Se tudo estiver correto, clique em 'Finalizar'.")
    
    # Resumo dos dados iniciais
    with st.expander("Resumo dos Dados Iniciais", expanded=True):
        dados = st.session_state.dados_iniciais
        
        # Contrato
        existe_contrato = dados.get('existe_contrato', 'Não')
        if existe_contrato == "Sim":
            texto_contrato = f"Sim (Nº {dados.get('n_contrato', 'N/A')})"
        else:
            texto_contrato = "Não"
        
        # Pedido
        existe_pedido = dados.get('existe_pedido', 'Não')
        if existe_pedido == "Sim":
            texto_pedido = f"Sim (Nº {dados.get('n_pedido', 'N/A')})"
        else:
            texto_pedido = "Não"
        
        col1, col2 = st.columns(2)
        with col1:
            st.write(f"**CNPJ:** {dados.get('cnpj', 'N/A')}")
            st.write(f"**Empresa:** {dados.get('empresa', 'N/A')}")
            st.write(f"**Advogado(a):** {dados.get('advogado', 'N/A')}")
            st.write(f"**Documento:** {dados.get('tipo_doc', 'N/A')}")
        
        with col2:
            st.write(f"**Data Emissão:** {dados.get('data_prevista').strftime('%d/%m/%Y') if dados.get('data_prevista') else 'N/A'}")
            st.write(f"**Nº medição:** {dados.get('n_medicao', 'N/A')}")
            st.write(f"**Contrato:** {texto_contrato}")
            st.write(f"**Pedido:** {texto_pedido}")
        
        st.write(f"**Breve descrição:** {dados.get('breve_desc', 'N/A')}")
    
    # Resumo das cobranças
    for i, cobranca in enumerate(st.session_state.dados_coletados):
        with st.expander(f"Resumo da Cobrança {i+1}", expanded=True):
            col1, col2 = st.columns(2)
            
            with col1:
                st.write(f"**Nº Espaider:** {cobranca.get('Nº Espaider', 'Não informado')}")
                st.write(f"**Projeto vinculado:** {cobranca.get('Projeto vinculado', 'Não informado')}")
                st.write(f"**Trecho:** {cobranca.get('Trecho', 'Não informado')}")
            
            with col2:
                st.write(f"**Matéria:** {cobranca.get('Matéria', 'N/A')}")
            
            st.markdown("---")
            
            # Cobrança Principal
            bloco_1 = cobranca.get('bloco_1', {})
            col1, col2 = st.columns(2)
            
            with col1:
                st.subheader("Cobrança Principal")
                st.write(f"**Tipo:** {bloco_1.get('tipo', 'N/A')}")
                st.write(f"**Valor:** R$ {bloco_1.get('valor', '0,00')}")
                st.write(f"**Texto Breve:** {bloco_1.get('texto_breve', 'N/A')}")
            
            # Cobrança Secundária se existir
            if 'bloco_2' in cobranca:
                bloco_2 = cobranca.get('bloco_2', {})
                with col2:
                    st.subheader("Cobrança Secundária")
                    st.write(f"**Tipo:** {bloco_2.get('tipo', 'N/A')}")
                    st.write(f"**Valor:** R$ {bloco_2.get('valor', '0,00')}")
                    st.write(f"**Texto Breve:** {bloco_2.get('texto_breve', 'N/A')}")
    
    # Botões de ação
    col1, col2, col3 = st.columns([1, 1, 1])
    
    with col1:
        if st.button("← Voltar e Corrigir Detalhes", use_container_width=True):
            navegar_para("detalhamento")
    
    with col3:
        if st.button("FINALIZAR E GERAR EXCEL", type="primary", use_container_width=True):
            finalizar_processo()

# ============================================================
# FUNÇÕES DE EXPORTAÇÃO EXCEL
# ============================================================

def formatar_valor_excel(valor_str):
    """Formata valor para Excel"""
    if not valor_str: 
        valor_str = "0,00"
    valor_limpo = re.sub(r'[^0-9,]', '', valor_str).replace(',', '.')
    try:
        return float(valor_limpo)
    except ValueError:
        return 0.0

def gerar_excel():
    """Gera arquivo Excel com todos os dados"""
    if not st.session_state.dados_iniciais or not st.session_state.dados_coletados:
        st.error("Não há dados de cobranças para gerar o arquivo Excel.")
        return None
    
    try:
        workbook = Workbook()
        
        # Planilha principal
        sheet_principal = workbook.active
        sheet_principal.title = "Medições"
        formatar_planilha_principal(sheet_principal)
        
        # Planilha BD
        sheet_bd = workbook.create_sheet(title="BD")
        formatar_planilha_bd(sheet_bd)
        
        workbook.active = sheet_principal
        
        # Proteção das planilhas
        sheet_principal.protection.password = 'SINAPSE4'
        sheet_principal.protection.sheet = True
        
        sheet_bd.protection.password = 'SINAPSE4'
        sheet_bd.protection.sheet = True
        sheet_bd.protection.autoFilter = False
        sheet_bd.sheet_state = 'hidden'
        
        workbook.security = WorkbookProtection(
            workbookPassword='SINAPSE4',
            lockStructure=True
        )
        
        # Salvar para buffer
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    except Exception as e:
        st.error(f"Erro ao gerar Excel: {e}")
        return None

def formatar_planilha_principal(sheet):
    """Formata a planilha principal"""
    font_header_azul = Font(name='Segoe UI', size=11, bold=True, color="FFFFFF")
    fill_header_azul = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    font_bold = Font(name='Segoe UI', size=10, bold=True)
    font_normal = Font(name='Segoe UI', size=10)
    
    alignment_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    alignment_right = Alignment(horizontal='right', vertical='center', wrap_text=True)
    alignment_center = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    sheet.sheet_view.showGridLines = False
    
    # Logo (placeholder)
    sheet['B2'] = "ISA Energia"
    sheet['B2'].font = Font(name='Segoe UI', size=14, bold=True)
    
    # Cabeçalho
    sheet.merge_cells('D2:G2')
    cell_med_juridica = sheet['D2']
    n_medicao_header = st.session_state.dados_iniciais.get('n_medicao', 'N/A')
    cell_med_juridica.value = f"Medição Jurídica Nº {n_medicao_header}"
    cell_med_juridica.font = Font(name='Segoe UI', size=16, bold=True)
    cell_med_juridica.alignment = alignment_center
    
    # Data
    cell_data_header = sheet['G4']
    cell_data_header.value = f"Data: {date.today().strftime('%d/%m/%Y')}"
    cell_data_header.font = font_normal
    cell_data_header.alignment = alignment_right
    
    # Dados Iniciais
    sheet.merge_cells('B5:G5')
    header_cell = sheet['B5']
    header_cell.value = "Dados Iniciais - Faturamento"
    header_cell.font = font_header_azul
    header_cell.fill = fill_header_azul
    header_cell.alignment = alignment_center
    
    for col_idx in range(2, 8):
        sheet.cell(row=5, column=col_idx).border = thin_border
    
    # Preencher dados iniciais
    dados = st.session_state.dados_iniciais
    dados_esquerda = [
        ("CNPJ do fornecedor:", dados['cnpj']),
        ("Empresa contratante:", dados['empresa']),
        ("Advogado responsável:", dados['advogado']),
        ("Tipo de documento de cobrança:", dados['tipo_doc']),
    ]
    
    current_row = 7
    for label, value in dados_esquerda:
        cell_label = sheet.cell(row=current_row, column=2)
        cell_label.value = label
        cell_label.font = font_bold
        cell_label.alignment = alignment_right
        
        sheet.merge_cells(f'C{current_row}:D{current_row}')
        cell_value = sheet.cell(row=current_row, column=3)
        cell_value.value = str(value) if value else ""
        cell_value.font = font_normal
        cell_value.alignment = alignment_left
        current_row += 1
    
    dados_direita = [
        ("Data prevista de emissão:", dados['data_prevista'].strftime('%d/%m/%Y')),
        ("Nº Contrato:", dados.get('n_contrato') if dados.get('existe_contrato') == "Sim" else "N/A"),
        ("Nº Pedido:", dados.get('n_pedido') if dados.get('existe_pedido') == "Sim" else "N/A")
    ]
    
    current_row = 7
    for label, value in dados_direita:
        cell_label = sheet.cell(row=current_row, column=5)
        cell_label.value = label
        cell_label.font = font_bold
        cell_label.alignment = alignment_right
        
        sheet.merge_cells(f'F{current_row}:G{current_row}')
        cell_value = sheet.cell(row=current_row, column=6)
        cell_value.value = str(value) if value else ""
        cell_value.font = font_normal
        cell_value.alignment = alignment_left
        current_row += 1
    
    # Detalhamento das cobranças
    current_row = 13
    
    sheet.merge_cells(f'B{current_row}:G{current_row}')
    header_cell_det = sheet.cell(row=current_row, column=2)
    header_cell_det.value = "Detalhamento das cobranças"
    header_cell_det.font = font_header_azul
    header_cell_det.fill = fill_header_azul
    header_cell_det.alignment = alignment_center
    
    current_row += 1
    
    # Cabeçalho da tabela de cobranças
    col_titles = {
        2: "Nº Espaider",
        3: "Projeto vinculado",
        4: "Trecho",
        5: "Tipo de Cobrança",
        6: "Matéria",
        7: "Valor",
        8: "Texto breve código serviço"
    }
    
    for col_idx, title in col_titles.items():
        cell = sheet.cell(row=current_row, column=col_idx)
        cell.value = title
        cell.font = font_header_azul
        cell.fill = fill_header_azul
        cell.border = thin_border
    
    # Dados das cobranças
    valor_total = 0.0
    
    for dados_cobranca in st.session_state.dados_coletados:
        bloco_1 = dados_cobranca.get('bloco_1', {})
        valor_bloco_1 = formatar_valor_excel(bloco_1.get("valor"))
        valor_total += valor_bloco_1
        
        current_row += 1
        dados_linha = {
            2: dados_cobranca.get("Nº Espaider", "N/A"),
            3: dados_cobranca.get("Projeto vinculado", "N/A"),
            4: dados_cobranca.get("Trecho", "N/A"),
            5: bloco_1.get("tipo"),
            6: bloco_1.get("materia"),
            7: valor_bloco_1,
            8: bloco_1.get("texto_breve")
        }
        
        for col_idx in range(2, 9):
            cell = sheet.cell(row=current_row, column=col_idx)
            cell.value = dados_linha.get(col_idx)
            cell.font = font_normal
            cell.border = thin_border
            cell.alignment = alignment_left
            
            if col_idx == 7:
                cell.number_format = 'R$ #,##0.00'
        
        # Cobrança secundária
        if 'bloco_2' in dados_cobranca:
            bloco_2 = dados_cobranca['bloco_2']
            valor_bloco_2 = formatar_valor_excel(bloco_2.get("valor"))
            valor_total += valor_bloco_2
            
            current_row += 1
            dados_linha = {
                2: dados_cobranca.get("Nº Espaider", "N/A"),
                3: dados_cobranca.get("Projeto vinculado", "N/A"),
                4: dados_cobranca.get("Trecho", "N/A"),
                5: bloco_2.get("tipo"),
                6: bloco_2.get("materia"),
                7: valor_bloco_2,
                8: bloco_2.get("texto_breve")
            }
            
            for col_idx in range(2, 9):
                cell = sheet.cell(row=current_row, column=col_idx)
                cell.value = dados_linha.get(col_idx)
                cell.font = font_normal
                cell.border = thin_border
                cell.alignment = alignment_left
                if col_idx == 7:
                    cell.number_format = 'R$ #,##0.00'
    
    # Total
    current_row += 1
    
    sheet.merge_cells(f'B{current_row}:E{current_row}')
    cell_total_label = sheet.cell(row=current_row, column=2)
    cell_total_label.value = "Valor total da cobrança"
    cell_total_label.font = font_bold
    cell_total_label.alignment = alignment_right
    
    cell_total_valor = sheet.cell(row=current_row, column=7)
    cell_total_valor.value = valor_total
    cell_total_valor.number_format = 'R$ #,##0.00'
    cell_total_valor.font = font_bold
    cell_total_valor.border = thin_border
    
    # Ajustar largura das colunas
    sheet.column_dimensions['A'].width = 2
    
    for col_idx in range(2, 9):
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        max_length = 0
        
        for cell in sheet[column_letter]:
            if cell.value:
                try:
                    cell_length = len(str(cell.value)) + 2
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
        
        adjusted_width = max(15, min(max_length, 50))
        
        if column_letter in ['C', 'D', 'E', 'F']:
            adjusted_width = max(25, min(max_length, 40))
        if column_letter == 'B':
            adjusted_width = max(20, min(max_length, 30))
        if column_letter == 'G':
            adjusted_width = 15
        if column_letter == 'H':
            adjusted_width = 25
        
        sheet.column_dimensions[column_letter].width = adjusted_width

def formatar_planilha_bd(sheet):
    """Formata a planilha BD"""
    header_iniciais = [
        "CNPJ Fornecedor", "Empresa Contratante", "Advogado(a) Responsável", "Tipo de Documento",
        "Data Emissão", "Qtd. Total Cobranças (Lançadas)", "Nº Contrato", "Nº Pedido", 
        "Nº medição (doc. fornecedor)"
    ]
    
    header_detalhe = [
        "Nº Cobrança", "Nº Espaider", "Projeto Vinculado", "Trecho", "Tipo de Cobrança",
        "Matéria Jurídica", "Valor (R$)", "Texto Breve Código Serviço"
    ]
    
    full_header = header_iniciais + header_detalhe
    sheet.append(full_header)
    
    font_header_bd = Font(name='Segoe UI', size=10, bold=True, color="FFFFFF")
    fill_header_bd = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    
    for cell in sheet[1]:
        cell.font = font_header_bd
        cell.fill = fill_header_bd
    
    sheet.auto_filter.ref = sheet.dimensions
    
    # Dados base
    dados = st.session_state.dados_iniciais
    total_cobrancas_lancadas = len(st.session_state.dados_coletados)
    
    dados_base = [
        dados['cnpj'], dados['empresa'], dados['advogado'], dados['tipo_doc'],
        dados['data_prevista'].strftime('%d/%m/%Y'), total_cobrancas_lancadas,
        dados.get('n_contrato') if dados.get('existe_contrato') == "Sim" else "",
        dados.get('n_pedido') if dados.get('existe_pedido') == "Sim" else "",
        dados.get('n_medicao', '')
    ]
    
    # Dados das cobranças
    for i, dados_cobranca in enumerate(st.session_state.dados_coletados):
        num_cobranca = i + 1
        n_espaider = dados_cobranca.get("Nº Espaider", "")
        projeto_vinculado = dados_cobranca.get("Projeto vinculado", "")
        trecho = dados_cobranca.get("Trecho", "")
        
        bloco_1 = dados_cobranca.get('bloco_1', {})
        dados_linha_detalhe_1 = [
            num_cobranca, n_espaider, projeto_vinculado, trecho,
            bloco_1.get("tipo"), bloco_1.get("materia"),
            formatar_valor_excel(bloco_1.get("valor")), bloco_1.get("texto_breve")
        ]
        sheet.append(dados_base + dados_linha_detalhe_1)
        
        if 'bloco_2' in dados_cobranca:
            bloco_2 = dados_cobranca['bloco_2']
            dados_linha_detalhe_2 = [
                num_cobranca, n_espaider, projeto_vinculado, trecho,
                bloco_2.get("tipo"), bloco_2.get("materia"),
                formatar_valor_excel(bloco_2.get("valor")), bloco_2.get("texto_breve")
            ]
            sheet.append(dados_base + dados_linha_detalhe_2)
    
    # Ajustar largura das colunas
    for col in sheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[column].width = adjusted_width

def finalizar_processo():
    """Finaliza o processo e gera o Excel"""
    if not st.session_state.dados_coletados:
        st.error("O processo não pode ser finalizado: não há detalhamentos válidos.")
        return
    
    st.session_state.is_finalizado = True
    
    # Gerar Excel
    buffer = gerar_excel()
    
    if buffer:
        # Download do arquivo
        nome_arquivo = f"Medicoes_Juridicas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        st.download_button(
            label="📥 Baixar Arquivo Excel",
            data=buffer,
            file_name=nome_arquivo,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        
        qtd = len(st.session_state.dados_coletados)
        if qtd == 1:
            mensagem = f"O detalhamento foi concluído e o arquivo de 1 cobrança foi gerado."
        else:
            mensagem = f"Os detalhamentos foram concluídos e o arquivo de {qtd} cobranças foi gerado."
        
        st.success(mensagem)

# ============================================================
# BARRA LATERAL
# ============================================================

def render_sidebar():
    """Renderiza a barra lateral com navegação e status"""
    with st.sidebar:
        st.markdown(
            f"<h2 style='color: {PRIMARY_BLUE}; text-align: center;'>Medições Jurídicas</h2>", 
            unsafe_allow_html=True
        )
        
        st.markdown("---")
        
        # Navegação
        pages = [
            ("🏠 Início", "inicio"),
            ("📊 Dados Iniciais", "dados"),
            ("🔍 Revisão", "revisao"),
            ("📝 Detalhamento", "detalhamento"),
            ("✅ Revisão Detalhada", "revisao_detalhada")
        ]
        
        current_page = st.session_state.current_page
        
        for label, page in pages:
            # Verificar se a página está habilitada
            disabled = False
            tooltip = ""
            
            if page == "revisao" and not st.session_state.is_dados_validado:
                disabled = True
                tooltip = "Valide os 'Dados Iniciais' para habilitar esta etapa"
            elif page == "detalhamento" and not st.session_state.is_revisao_concluida:
                disabled = True
                tooltip = "Confirme a 'Revisão' para habilitar esta etapa"
            elif page == "revisao_detalhada" and not st.session_state.is_detalhamento_validado:
                disabled = True
                tooltip = "Valide o 'Detalhamento' para habilitar esta etapa"
            
            if disabled:
                st.button(
                    label, 
                    disabled=True,
                    help=tooltip,
                    use_container_width=True
                )
            else:
                if st.button(
                    label,
                    type="primary" if current_page == page else "secondary",
                    use_container_width=True
                ):
                    navegar_para(page)
        
        st.markdown("---")
        
        # Status do processo
        st.markdown("**ETAPAS DO PROCESSO**")
        
        etapas = [
            ("1. Dados Iniciais", st.session_state.is_dados_validado),
            ("2. Revisão", st.session_state.is_revisao_concluida),
            ("3. Detalhamento", st.session_state.is_detalhamento_validado),
            ("4. Revisão Detalhada", st.session_state.is_detalhamento_validado),
            ("5. Geração do Arquivo", st.session_state.is_finalizado)
        ]
        
        for etapa, concluida in etapas:
            if concluida:
                st.markdown(f"✅ {etapa}")
            else:
                st.markdown(f"◯ {etapa}")
        
        st.markdown("---")
        st.markdown("*© Equipe de Desenvolvimento*")

# ============================================================
# APLICAÇÃO PRINCIPAL
# ============================================================

def main():
    """Função principal da aplicação"""
    
    # Configuração da página
    st.set_page_config(
        page_title="Sistema de Medições Jurídicas",
        page_icon="⚖️",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Inicializar estados
    initialize_session_state()
    
    # CSS personalizado
    st.markdown(f"""
        <style>
        .main .block-container {{
            padding-top: 2rem;
        }}
        .stButton>button {{
            width: 100%;
        }}
        .stExpander {{
            background-color: {LIGHT_GRAY_BG};
        }}
        </style>
    """, unsafe_allow_html=True)
    
    # Barra lateral
    render_sidebar()
    
    # Conteúdo principal baseado na página atual
    current_page = st.session_state.current_page
    
    if current_page == "inicio":
        pagina_inicio()
    elif current_page == "dados":
        pagina_dados()
    elif current_page == "revisao":
        pagina_revisao()
    elif current_page == "detalhamento":
        pagina_detalhamento()
    elif current_page == "revisao_detalhada":
        pagina_revisao_detalhada()

if __name__ == "__main__":
    main()
